			# Currently configured for 18-way split.
			# Now also writes out prop. data to an Excel file.
			# Extracts the same data as table_pairwise.py, but converts the pairwise OGs to % total for each group.
			# Now modified for ease of use in R. Proportional data is correctly ordered, and within-group data set to 0.

from openpyxl import Workbook
from openpyxl import load_workbook
import re
import group
import glob

# correct_order and group_names can be switched as needed.
# Also remember to configure n in chunks(l, n) to correspond to number of groups in group_names.
# correct_order is correct for proportion of total genome shared.
output = open("vector_ordered_propdata_18.txt", "w")
correct_order_12 = ["SAR", "Haptista", "Archaeplastida", "Ancyromonadida", "Telonemids", "Obazoa", "Discoba", "Collodictyonids", "Cryptista", "Amoebozoa", "Metamonads", "Malawimonadidae"]
correct_order_15 = ["Telonemids", "Haptista", "SAR", "Atwista", "Archaeplastida", "Ancyromonadida", "Obazoa", "Discoba", "Collodictyonids", "Cryptista", "Amoebozoa", "Metamonads", "Hemimastigophora", "Apusomonada", "Malawimonadidae"]
correct_order_includingown_18 = ["Alveolata", "Stramenopiles", "Archaeplastida", "Obazoa", "Telonemids", "Centrohelids", "Ancyromonadida", "Discoba", "Rhizaria", "Cryptista", "Atwista", "Haptophyta", "Collodictyonids", "Amoebozoa", "Metamonads", "Hemimastigophora", "Apusomonada", "Malawimonadidae"]

# Choose the necessary group list from the group module.
alt_groups = sorted(group.alt_groups_18())
group_names = alt_groups
data = []

for name in group_names:
	to_parse = glob.glob("/mnt/c/Users/scamb/Documents/uob_msc/Genome_data/OG_arb-fal/new_outputs/*.txt")
	for file in to_parse:
		filename = re.search(r"([A-Z]\w*)_([A-Z]\w*)_output.txt$", file)
		if filename:
			if filename.group(1) not in group_names:		# Skips 'non-target' files in new_outputs.
				pass
			elif filename.group(2) not in group_names:		# Does the same as above (Since name order in filename can vary).
				pass
			elif filename.group(1) == name:
				shared_og = group.parse_OG(file)
				data.append(shared_og)
			elif filename.group(2) == name:
				shared_og = group.parse_OG(file)
				data.append(shared_og)

def chunks(l, n):
	for i in range(0, len(l), n):			# The 3rd argument n is the stepping distance that i jumps after each iteration.
		yield l[i:i + n]			# Yield is here used in place of return, since it returns a sequence of lists (rather than terminating after the first list is returned).

chunkdata = list(chunks(data, 18))			# n MUST be configured depending on number in group_names.
totalOG = group.parse_total()				# Retrieving totals dictionary.

propdata = []
i = 0							# Necessary to include counter for loop since the "groupth" iteration not compatible with both dict and list types.

for group in group_names:				# Calculating genome similarity
	totals = totalOG[group]				# Total values are accessed.
							# propdata.append(group) can be added here if desired (group name is appended corresponding to the relevant data in chunkdata).
	for element in chunkdata[i]:
		proportion = round(int(element) / int(totals) * 100, 2)		# Returns % similarity of genome for each pairwise comparison.
		propdata.append(proportion)
	i += 1

xl_data = []

for eugroup in correct_order_includingown_18:									# Choose list as appropriate.
	j = 0
	for pos, name in enumerate(group_names):
		group_data = propdata[j:j + len(group_names)]						# Captures the data corresponding to the group.
		j += len(group_names)									# j ensures the data index range goes up by the len of group_names.
		group_data = [float(number) for number in group_data]					# Converts to a list of floats (for proportion)..
		if name == eugroup:
			for index, x in enumerate(group_data):						# Loop comparing indices of group_names and group_data.
				if index == pos:
					group_data.insert(0, eugroup)					# insert() adds in the group name at any index.
					xl_data.append(group_data)					# Builds up list ready to be written out to Excel.
#					group_data[index] = 0						# Sets own group values to 0.
					translation_table = dict.fromkeys(map(ord, "\w+[',]"), None) 	# This is what I use to strip lists of their punctuation.
					fdata = str(group_data).translate(translation_table)
					outputWrite = output.write(f"{fdata} ")

# xl_data takes proportional data before own group data is set to zero.
# But - This still somehow sets values to zero, so I must switch off for outputting to Excel. Do not understand...
# This data is now written out to Excel - the same file storing non-proportional data.
wb = load_workbook(filename = "/mnt/c/Users/scamb/Documents/uob_msc/Genome_data/data2go/Pairwise_proportional_18.xlsx")
new_ws = wb.create_sheet("Proportional data")

new_ws.append(group_names)

for row in xl_data:
	new_ws.append(row)

wb.save("Pairwise_proportional_18_copy.xlsx")

output.close()
